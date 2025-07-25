import pandas as pd
from openpyxl import Workbook, load_workbook
import os

csv_input = "MD/cincinnati_childrens/raw/cincinnati_childrens_hospital.csv"
excel_output = "MD/cincinnati_childrens/raw/cincinnati_childrens_hospital.xlsx"
output_dir = "MD/cincinnati_childrens/raw_final"

os.makedirs(output_dir, exist_ok=True)

df = pd.read_csv(csv_input, dtype=str)  # reading all values as strings to prevent auto-formatting in excel
df.to_excel(excel_output, index=False)

# Desired filenames for each block
block_names = [
    "raw_subject_characteristics",
    "raw_timing",
    "raw_family_medical_history",
    "raw_demographics",
    "raw_medical_history",
    "raw_biometrics",
    "raw_genetic_analysis",
    "raw_testing",
    "raw_disease_characteristics",
    "raw_medication"
]

wb_in = load_workbook(filename=excel_output)
ws_in = wb_in.active

header = [cell.value.strip() if isinstance(cell.value, str) else cell.value for cell in ws_in[1]]

data_rows = list(ws_in.iter_rows(min_row=2, values_only=True))

complete_indices = [i for i, col in enumerate(header) if isinstance(col, str) and col.strip().startswith("Complete")]

try:
    id_index = header.index("Honest Broker Subject ID")
except ValueError:
    raise ValueError("❌ 'Honest Broker Subject ID' not found.")

start_idx = id_index + 1

for block_num, end_idx in enumerate(complete_indices, start=0):  
    if block_num >= len(block_names):
        print(f"⚠️ Skipping block {block_num+1} - No matching name in block_names list.")
        continue

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Sheet1"

    col_indices = [id_index] + list(range(start_idx, end_idx + 1))
    block_header = [header[i] for i in col_indices]
    ws_out.append(block_header)

    for row in data_rows:
        block_row = [row[i] if i < len(row) else None for i in col_indices]
        ws_out.append(block_row)

    output_name = block_names[block_num]
    output_path = f"{output_dir}/{output_name}.xlsx"
    wb_out.save(output_path)
    print(f"✓ Saved '{output_name}' to {output_path}")

    # preparing start index for next block
    start_idx = end_idx + 1
